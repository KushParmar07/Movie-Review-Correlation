from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
import openpyxl
import yaml

def analyze_sentiment(sheet):
    sentiment_scores = []
    analyzer = SentimentIntensityAnalyzer()
    try:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    sentiment = analyzer.polarity_scores(cell.value)
                    sentiment_scores.append(sentiment['compound'])
    except:
        pass

    return sentiment_scores

if __name__ == "__main__":
    workbook = openpyxl.load_workbook('Movie_Reviews.xlsx')
    try:
        for sheet_name in workbook.sheetnames:
            print(f"Analyzing {sheet_name}...")
            sheet = workbook[sheet_name]
            sentiment_scores = analyze_sentiment(sheet)
            with open('sentiment_scores.txt', 'a') as f:
                if sentiment_scores:
                    f.write(f"{sheet_name}: {sum(sentiment_scores) / len(sentiment_scores)}\n")

    except:
        pass




